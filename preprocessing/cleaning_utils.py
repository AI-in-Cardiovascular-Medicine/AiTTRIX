import numpy as np
import pandas as pd
from missforest import MissForest
from openpyxl import load_workbook
import sys

class Tee:
    def __init__(self, filename, mode="w"):
        self.terminal = sys.stdout
        self.log = open(filename, mode)
        print(f"Save preprocessing to {filename}")

    def write(self, message):
        self.terminal.write(message)   # print to terminal
        self.log.write(message)        # also write to file

    def flush(self):  # needed for interactive use
        self.terminal.flush()
        self.log.flush()

def remove_features_with_many_nan(data, nan_threshold, return_drop_features=False, verbose=True):
    """
    Remove columns from a DataFrame that have a high proportion of missing values.

    Args:
        data (pd.DataFrame): Input DataFrame containing features.
        nan_threshold (float): Maximum allowed proportion of NaN values per column (between 0 and 1).
                               Columns exceeding this threshold are dropped.
        return_drop_features (bool, optional): If True, also return the list of dropped columns. Defaults to False.
        verbose (bool, optional): If True, prints the names of columns removed. Defaults to True.

    Returns:
        pd.DataFrame: DataFrame with columns exceeding the NaN threshold removed.
        list, optional: List of dropped column names (only returned if `return_drop_features=True`).
    """
    null_perc = data.isna().sum() / len(data)
    to_drop = null_perc[null_perc > nan_threshold].index.values.tolist()
    if len(to_drop) > 0 and verbose:
        print(f"Removing {len(to_drop)} features (more than {nan_threshold*100}% of nans):\n\t{to_drop}")
    data = data.drop(columns=to_drop)
    if return_drop_features:
        return data, to_drop
    else:
        return data


def remove_patients_with_many_nan(data, nan_threshold):
    """
    Remove patients (rows) from a DataFrame that have a high proportion of missing values.

    Args:
        data (pd.DataFrame): Input DataFrame containing patient data.
        nan_threshold (float): Maximum allowed proportion of NaN values per row (between 0 and 1).
                               Rows exceeding this threshold are dropped.

    Returns:
        pd.DataFrame: DataFrame with patients exceeding the NaN threshold removed.
    """
    many_na_mask = data.isna().sum(axis=1) / len(data.columns) > nan_threshold
    if sum(many_na_mask) > 0:
        print(f"Removing {sum(many_na_mask)} patients with more than {nan_threshold*100}% of nans")
    data = data.drop(data[many_na_mask].index)
    return data


def remove_highly_correlated_features(data, outcome, corr_threshold=0.9):
    """
    Remove highly correlated features from a DataFrame based on a correlation threshold.

    Args:
        data (pd.DataFrame): Input DataFrame containing features and the outcome column.
        outcome (str): Name of the outcome column to preserve.
        corr_threshold (float, optional): Maximum allowed correlation between features (default=0.9).
                                          Features with higher correlation are dropped.

    Returns:
        pd.DataFrame: DataFrame with highly correlated features removed and the outcome column preserved.
    """
    y = data[outcome]
    data = data.drop(columns=outcome)
    corr_matrix = data.corr(numeric_only=True)
    importances = data.corrwith(y, axis=0, numeric_only=True).abs()
    importances = importances.sort_values(ascending=False)
    corr_matrix = corr_matrix.reindex(index=importances.index, columns=importances.index).abs()
    upper_triangle = corr_matrix.where(np.triu(np.ones(corr_matrix.shape), k=1).astype(bool))
    to_drop = [col for col in upper_triangle.columns if any(upper_triangle[col] > corr_threshold)]
    if len(to_drop) > 0:
        print(f"Removing {len(to_drop)} highly correlated features (corr > {corr_threshold}): {to_drop}")
    data = data.drop(columns=to_drop)
    data[outcome] = y
    return data

def remove_binaries_not_populated(data, binary_threshold=0.01, return_drop_features=False, verbose=True):
    """
    Remove binary features that are too sparsely or almost fully populated.

    Args:
        data (pd.DataFrame): Input DataFrame containing features.
        binary_threshold (float, optional): Minimum or maximum fraction for a binary feature to be retained
                                            (default=0.01). Features with values < threshold or > 1-threshold are dropped.
        return_drop_features (bool, optional): If True, also return the list of dropped features. Defaults to False.
        verbose (bool, optional): If True, prints information about removed features. Defaults to True.

    Returns:
        pd.DataFrame: DataFrame with low-populated binary features removed.
        list, optional: List of dropped binary feature names (only returned if `return_drop_features=True`).
    """
    binary_features = []
    for col in data.columns:
        values = data.loc[:, col].dropna()
        if values.ndim == 1:  # skip if it's somehow a DataFrame
            if set(values.unique()) <= {0, 1, "0", "1"}:
                binary_features.append(col)
    data[binary_features] = data[binary_features].apply(pd.to_numeric, errors="coerce")
    binary_frac = data[binary_features].sum() / len(data)
    to_drop = binary_frac[(binary_frac < binary_threshold) | (binary_frac > 1 - binary_threshold)]
    if len(to_drop) > 0 and verbose:
        summary = pd.DataFrame({
            "percentage": to_drop * 100,
            "count": (data[binary_features].sum())[to_drop.index]
        })
        print(
            f"Removing {len(to_drop)} low populated binary features "
            f"(% < {binary_threshold * 100} or % > {100 - binary_threshold * 100}):"
        )
        print(summary)
    data = data.drop(columns=to_drop.index.values)
    if return_drop_features:
        return data, to_drop.index.values.tolist()
    else:
        return data


def remove_0_variance_features(data, return_drop_features=False, verbose=True):
    """
    Remove features with zero variance from a DataFrame.

    Args:
        data (pd.DataFrame): Input DataFrame containing features.
        return_drop_features (bool, optional): If True, also return the list of dropped features. Defaults to False.
        verbose (bool, optional): If True, prints information about removed features. Defaults to True.

    Returns:
        pd.DataFrame: DataFrame with zero-variance features removed.
        list, optional: List of removed feature names (only returned if `return_drop_features=True`).
    """
    data_var = data.var(numeric_only=True)
    data_var_0 = data_var[data_var == 0].index.values
    if len(data_var_0) > 0 and verbose:
        print(f"Removing {len(data_var_0)} features with 0 variance:\n\t{data_var_0}")
    data = data.drop(columns=data_var_0)
    if return_drop_features:
        return data, data_var_0.tolist()
    else:
        return data


def remove_patients_without_outcome(data, time_column, event_column):
    """
    Remove patients (rows) from a DataFrame that are missing time-to-event or event labels.

    Args:
        data (pd.DataFrame): Input DataFrame containing patient data.
        time_column (str): Name of the column representing time-to-event.
        event_column (str): Name of the column representing the event indicator.

    Returns:
        pd.DataFrame: DataFrame with patients missing time or event information removed.
    """
    indices_no_outcome = data[(data[time_column].isna()) | (data[event_column].isna())].index
    if len(indices_no_outcome) > 0:
        print(f"Removing {len(indices_no_outcome)} patients without time to event or event label")
    data = data.drop(indices_no_outcome)
    return data


def impute_data(train, test=None, max_categories=10):
    count_unique = train.nunique()
    categorical = list(count_unique[count_unique < max_categories].index)
    if len(categorical) == 0:
        categorical = None
    imputer = MissForest(categorical=categorical)
    train_imputed = train.copy()
    if train.isna().sum().sum() > 0:
        imputer.fit(train)
        imp_train = imputer.transform(train)
        train_imputed = pd.DataFrame(imp_train, index=train.index, columns=train.columns)
    if test is not None and test.isna().sum().sum() > 0:
        imp_test = imputer.transform(test)
        test_imputed = pd.DataFrame(imp_test, index=test.index, columns=test.columns)
        return train_imputed, test_imputed
    else:
        return train_imputed


def get_colored_columns(file_path, sheet_name):
    # Read file and select sheet
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    # Get colored columns
    colored_columns = []
    for col in sheet.iter_cols(min_row=1, max_row=1):  # Only check the header row
        cell = col[0]  # Get the first cell in the column
        if cell.fill and cell.fill.start_color.rgb != "00000000":  # Check if filled
            colored_columns.append((cell.value, cell.fill.start_color.rgb))
    return colored_columns


def set_outliers_to_nan(df, max_categories=5, exclude_columns=None):
    """
    Set outliers to NaN values for continuous variables in a DataFrame using the IQR method.

    Args:
        df (pd.DataFrame): Input DataFrame containing features.
        max_categories (int, optional): Maximum number of unique values to consider a column as categorical
                                        (default=5). Columns with fewer unique values are treated as categorical.
        exclude_columns (list, optional): List of columns to always treat as categorical/excluded from outlier detection.

    Returns:
        pd.DataFrame: DataFrame where outliers in continuous variables are replaced with NaN.
    """
    df = df.copy()
    unique_values = df.nunique()
    categorical = list(unique_values[unique_values <= max_categories].index)
    categorical = categorical + exclude_columns if exclude_columns is not None else categorical
    continuous = [col for col in df.columns if col not in categorical]
    q1 = df[continuous].quantile(0.01)
    q3 = df[continuous].quantile(0.99)
    iqr = q3 - q1
    mask = (df[continuous] > q3 + iqr) | (df[continuous] < q1 - iqr)
    for col in continuous:
        if np.sum(mask[col]) > 0:
            print(f"Set {np.sum(mask[col])} nans for feature {col}, values {df.loc[mask[col], col].values}")
    df[mask] = np.nan
    return df


def get_continuous_categorical_features(df, max_categories=10, separate_binary=True):
    """
    Split DataFrame columns into continuous, categorical, and optionally binary features.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    max_categories : int, default=10
        Maximum number of unique values for a column to be considered categorical.
    separate_binary : bool, default=True
        Whether to treat binary columns (2 unique values) separately.

    Returns
    -------
    continuous_cols : list
        Columns considered continuous.
    categorical_cols : list
        Columns considered categorical.
    binary_cols : list, optional
        Columns with exactly 2 unique values (returned only if separate_binary=True).
    """
    n_unique = df.nunique()
    if separate_binary:
        binary_cols = n_unique[n_unique == 2].index.tolist()
        categorical_cols = n_unique[(n_unique > 2) & (n_unique < max_categories)].index.tolist()
        continuous_cols = [col for col in df.columns if col not in binary_cols + categorical_cols]
        return continuous_cols, categorical_cols, binary_cols
    else:
        categorical_cols = n_unique[n_unique < max_categories].index.tolist()
        continuous_cols = [col for col in df.columns if col not in categorical_cols]
        return continuous_cols, categorical_cols
